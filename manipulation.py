# -*- coding: utf-8 -*-
#Employee Work Sheet Automation

# Make connection to google sheets 
from google.colab import auth
auth.authenticate_user()
# Library imports 
import gspread
from oauth2client.client import GoogleCredentials
import pandas as pd
import datetime
from google.colab import files
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# Authentication of google account 
gc = gspread.authorize(GoogleCredentials.get_application_default())

# Open Google Sheet (sheet 1) based on name
worksheet = gc.open('Time Sheet Entries  (Responses)').sheet1

# Get a list of rows
rows = worksheet.get_all_values()
# Convert list of rows to a DataFrame
employeeData = pd.DataFrame.from_records(rows[1:], columns=rows[0])
# View column names of dataframe
employeeData.columns

# The columns repeate names due to google form so create unique identifier for each repeated column
identifier = employeeData.columns.to_series().groupby(level=0).transform('cumcount')
# Rename columns with the new identifiers
employeeData.columns = employeeData.columns.astype('string') + '_' + identifier.astype('string')

# Authentication of google account 
gc = gspread.authorize(GoogleCredentials.get_application_default())

# Open Google Sheet (sheet 1) based on name
worksheet = gc.open('subcontractors').sheet1

# Get a list of rows
rows1 = worksheet.get_all_values()
# Convert list of rows to a DataFrame
employeeTable = pd.DataFrame.from_records(rows1[1:], columns=rows1[0])

def createAppend(date, name, rate, row, num:int, is_All:bool):
  if is_All:
    append_Dict = {'Timestamp':date, 'Person':name, 'LT#':row[f'Lot Number_{num}'],
                   'Project':row[f'Project Name_{num}'],
                   'Board':row[f'Board_{num}'], 'Tape':row[f'Tape _{num}'],
                   'Prepwork':row[f'Tape _{num}'],
                   'Twelve':row[f'Number of sheets (12 inches)_{num}'],
                   'Ten':row[f'Number of sheets (10 inches)_{num}'],
                   'Nine':row[f'Number of sheets (9 inches)_{num}'],
                   'Eight':row[f'Number of sheets (8 inches)_{num}'],
                   'SQF':000, 'Rate':rate, 'Total':000,
                   'Notes':row[f'Additional Notes_{num}']}
  else:
    append_Dict = {'LT#':row[f'Lot Number_{num}'],
                   'Project':row[f'Project Name_{num}'],
                   'Board':row[f'Board_{num}'], 'Tape':row[f'Tape _{num}'],
                   'Prepwork':row[f'Tape _{num}'],
                   'Twelve':row[f'Number of sheets (12 inches)_{num}'],
                   'Ten':row[f'Number of sheets (10 inches)_{num}'],
                   'Nine':row[f'Number of sheets (9 inches)_{num}'],
                   'Eight':row[f'Number of sheets (8 inches)_{num}'],
                   'SQF':000, 'Rate':rate, 'Total':000,
                   'Notes':row[f'Additional Notes_{num}']}
  return append_Dict

def addSummaryofWeek(df, week_of:int):

  # Create summary of week dataframe
  columns = ['Timestamp', 'Person', 'LT#', 'Project', 'Board', 'Tape', 'Twelve', 'Ten', 
             'Nine', 'Eight', 'SQF', 'Rate', 'Total', 'Notes']
  all_DataFrame = pd.DataFrame(columns=columns)

  # Keep appending rows to dataframe for each employee and each project
  for index, row in df.loc[:].iterrows():
    date = pd.to_datetime(row['Timestamp_0'], format='%Y%m%d', errors='ignore')
    name = row['First Name_0'] + " " + row['Last Name_0']

    # Get employee rate from employee table
    employeeDF = employeeTable.loc[(employeeTable['First Name'] == row['First Name_0']) & (employeeTable['Last Name'] == row['Last Name_0'])]
    rate = employeeDF.iloc[0]['Rate']

    for i in range(0,20):
      if row[f'Lot Number_{i}']!='':
        # Function call to row dictionary
        data_ToAppend=createAppend(date, name, rate, row, i, True)
        # Calculating SQF 
        num = ['Twelve', 'Ten', 'Nine', 'Eight']
        for i in num: 
          data_ToAppend[i] = int(data_ToAppend[i])
        data_ToAppend["A"] = data_ToAppend["Twelve"] * 48
        data_ToAppend["B"] = data_ToAppend["Ten"] * 40
        data_ToAppend["C"] = data_ToAppend["Nine"] * 36
        data_ToAppend["D"] = data_ToAppend["Eight"] * 32
        listing = ['A', 'B', 'C', 'D']
        sum = 0
        for character in listing:
          for key, value in data_ToAppend.items():
            if key == character:
              sum = value + sum
        data_ToAppend["SQF"] = sum
        keys_to_remove = ["A", "B", "C", "D"]
        for key in keys_to_remove:
          del data_ToAppend[key]
        
        # Calculating Total per project
        data_ToAppend["Total"] = float(data_ToAppend["SQF"]) * float(data_ToAppend["Rate"])
        res = dict()
        for k, v in data_ToAppend.items():
          if k == 'Total':
            # Rounding to 2 using round()
            data_ToAppend[k] = round(data_ToAppend[k], 2)
        all_DataFrame = all_DataFrame.append(data_ToAppend, ignore_index=True)

  all_DataFrame[['Date','Time']] = all_DataFrame['Timestamp'].str.split(' ',expand=True)
  all_DataFrame = all_DataFrame.drop(columns=['Timestamp', 'Time'])

  # Add final sheet to excel for all employee information 
  filepath = f'/content/drive/My Drive/Work_Sheet_{week_of}.xlsx'
  print('The file path of excel sheet created for this week is:', filepath)
  all_DataFrame = all_DataFrame[['Date', 'Person', 'LT#', 'Project', 'Board', 
                                 'Tape', 'Twelve', 'Ten', 'Nine', 'Eight', 
                                 'Prepwork','SQF', 'Rate', 'Total','Notes']]
  # Create excel with all employees entered being the first sheet
  #all_DataFrame.loc[:, cols] = all_DataFrame[cols].astype(float).applymap('${:,.2f}'.format)
  all_DataFrame.to_excel(filepath, sheet_name = 'All_Employees', index=False)

# Function to create excel work sheet - Main function 
def createWorkSheet(df, employeeDF, week_of:int):

  addSummaryofWeek(df, week_of)

  # Keep track of excel sheet names so there is no repetition 
  sheet_names = []
  num = ['Twelve', 'Ten', 'Nine', 'Eight']

  # Iterate through dataframe - each employee is a new sheet
  for index, row in employeeData.loc[:].iterrows():
    # Create time sheet dataframe
    columns = ['LT#', 'Project', 'Board', 'Tape', 'Twelve', 'Ten', 'Nine', 
               'Eight', 'SQF', 'Rate', 'Total', 'Notes']
    ind_DataFrame = pd.DataFrame(columns=columns)
    
    # Fix date format
    date = pd.to_datetime(row['Timestamp_0'], format='%Y%m%d', errors='ignore')
    file_date = str(date).split(' ')[0]
    file_date = file_date.replace("/", "_")
    # Get employee name 
    name = row['First Name_0'] + "_" + row['Last Name_0'] 

    # Get employee rate from employee table
    employeeDF = employeeTable.loc[(employeeTable['First Name'] == row['First Name_0']) & (employeeTable['Last Name'] == row['Last Name_0'])]
    rate = employeeDF.iloc[0]['Rate']

    if f'{name}_Sheet' in sheet_names:
      already_df = pd.read_excel('/content/drive/My Drive/Work_Sheet_20210417.xlsx', f'{name}_Sheet')
      already_df = already_df.loc[(already_df['LT#'] != 'Total') & (already_df['LT#'] != 'Extras:') 
      & (already_df['LT#'] != 'Deductions:') & (already_df['LT#'] != 'Absolute Total:')
      & (already_df['LT#'] != 'Total Extras:') & (already_df['LT#'] != 'Total Deductions:')]
      
      for i in range(0,20):
        if row[f'Lot Number_{i}']!='':
          # Function call to row dictionary
          data_ToAppend=createAppend(date, name, rate, row, i, False)
          
          # Calculating SQF
          for i in num: 
            data_ToAppend[i] = int(data_ToAppend[i])
          data_ToAppend["A"] = data_ToAppend["Twelve"] * 48
          data_ToAppend["B"] = data_ToAppend["Ten"] * 40
          data_ToAppend["C"] = data_ToAppend["Nine"] * 36
          data_ToAppend["D"] = data_ToAppend["Eight"] * 32
          listing = ['A', 'B', 'C', 'D']
          sum = 0
          for character in listing:
            for key, value in data_ToAppend.items():
              if key == character:
                sum = value + sum
          data_ToAppend["SQF"] = sum
          keys_to_remove = ["A", "B", "C", "D"]
          for key in keys_to_remove:
            del data_ToAppend[key]

          # Calculating Total per project
          data_ToAppend["Total"] = float(data_ToAppend["SQF"]) * float(data_ToAppend["Rate"])
          res = dict()
          for k, v in data_ToAppend.items():
            if k == 'Total':
              # rounding to 2 using round()
              data_ToAppend[k] = round(data_ToAppend[k], 2)
          already_df = already_df.append(data_ToAppend, ignore_index=True)

      # Final row is a total of SQF and Money for employee that week 
      already_df.SQF = already_df.SQF.astype(float)
      already_df.Total = already_df.Total.astype(float)
      # Get the total sum of SQF and Total
      SQF_sum = already_df.SQF.sum()
      Total_sum = already_df.Total.sum()
      append_Dict = {'LT#':'Total', 'Project':'-', 'Board':'-', 'Tape':'-',
                    'Twelve':'-', 'Ten':'-', 'Nine':'-', 'Eight':'-', 
                    'SQF':SQF_sum, 'Rate':'-', 
                    'Total':Total_sum, 'Notes':'-'}
      already_df = already_df.append(append_Dict, ignore_index=True)
      for i in range(0,5):
        append_Dict = {'LT#':'Extras:', 'Project':'Enter Text Here', 
                      'Board':'-', 'Tape':'-',
                      'Twelve':'-', 'Ten':'-', 'Nine':'-', 'Eight':'-', 
                      'SQF':'-', 'Rate':'-', 
                      'Total':'-', 'Notes':'-'}
        already_df = already_df.append(append_Dict, ignore_index=True)
      for i in range(0,5):
        append_Dict = {'LT#':'Deductions:', 'Project':'Enter Text Here', 
                      'Board':'-', 'Tape':'-', 'Twelve':'-', 'Ten':'-', 
                      'Nine':'-', 'Eight':'-', 'SQF':'-', 'Rate':'-', 
                      'Total':'-', 'Notes':'-'}
        already_df = already_df.append(append_Dict, ignore_index=True)
      append_Dict = {'LT#':'Absolute Total:'}
      already_df = already_df.append(append_Dict, ignore_index=True)
      already_df = already_df[['LT#', 'Project', 'Board', 'Tape', 'Twelve', 
                               'Ten', 'Nine', 'Eight', 'Prepwork','SQF', 
                               'Rate', 'Total','Notes']]

      filepath = f'/content/drive/My Drive/Work_Sheet_{week_of}.xlsx'

      # Modify sheet from excel for subsequent employee that already existed 
      book = load_workbook(filepath)
      writer = pd.ExcelWriter(filepath, engine='openpyxl')
      writer.book = book
      writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
      print(f'Pre-existing excel sheet named: {name}_Sheet has been adjusted.')
      already_df.to_excel(writer, f'{name}_Sheet', index=False)
      writer.save()

    else:
      # Depending on how many garages were made, keep appending row to dataframe 
      for i in range(0,20):

        if row[f'Lot Number_{i}']!='':
          # Function call to row dictionary
          data_ToAppend=createAppend(date, name, rate, row, i, False)

          # Calculating SQF
          for i in num: 
            data_ToAppend[i] = int(data_ToAppend[i])
          data_ToAppend["A"] = data_ToAppend["Twelve"] * 48
          data_ToAppend["B"] = data_ToAppend["Ten"] * 40
          data_ToAppend["C"] = data_ToAppend["Nine"] * 36
          data_ToAppend["D"] = data_ToAppend["Eight"] * 32
          listing = ['A', 'B', 'C', 'D']
          sum = 0
          for character in listing:
            for key, value in data_ToAppend.items():
              if key == character:
                sum = value + sum
          data_ToAppend["SQF"] = sum
          keys_to_remove = ["A", "B", "C", "D"]
          for key in keys_to_remove:
            del data_ToAppend[key]
          
          # Calculating Total per project 
          data_ToAppend["Total"] = float(data_ToAppend["SQF"]) * float(data_ToAppend["Rate"])
          res = dict()
          for k, v in data_ToAppend.items():
            if k == 'Total':
              # rounding to 2 using round()
              data_ToAppend[k] = round(data_ToAppend[k], 2)
          ind_DataFrame = ind_DataFrame.append(data_ToAppend, ignore_index=True)

      

      # Final row is a total of SQF and Money for employee that week 
      ind_DataFrame.SQF = ind_DataFrame.SQF.astype(float)
      ind_DataFrame.Total = ind_DataFrame.Total.astype(float)
      # Get the total sum of SQF and Total
      SQF_sum = ind_DataFrame.SQF.sum()
      Total_sum = ind_DataFrame.Total.sum()
      append_Dict = {'LT#':'Total', 'Project':'-', 'Board':'-', 'Tape':'-',
                   'Twelve':'-', 'Ten':'-', 'Nine':'-', 'Eight':'-', 
                   'SQF':SQF_sum, 'Rate':'-', 'Total':Total_sum, 
                   'Notes':'-', 'Prepwork':'-'}
      ind_DataFrame = ind_DataFrame.append(append_Dict, ignore_index=True)
      for i in range(0,5):
        append_Dict = {'LT#':'Extras:', 'Project':'Enter Text Here', 
                      'Board':'-', 'Tape':'-','Twelve':'-', 'Ten':'-', 
                      'Nine':'-', 'Eight':'-', 'SQF':'-', 'Rate':'-', 
                      'Total':'-', 'Notes':'-'}
        ind_DataFrame = ind_DataFrame.append(append_Dict, ignore_index=True)
      for i in range(0,5):
        append_Dict = {'LT#':'Deductions:', 'Project':'Enter Text Here', 
                      'Board':'-', 'Tape':'-', 'Twelve':'-', 'Ten':'-', 
                      'Nine':'-', 'Eight':'-', 'SQF':'-', 'Rate':'-', 
                      'Total':'-', 'Notes':'-'}
        ind_DataFrame = ind_DataFrame.append(append_Dict, ignore_index=True)
      append_Dict = {'LT#':'Absolute Total:'}
      ind_DataFrame = ind_DataFrame.append(append_Dict, ignore_index=True)
      ind_DataFrame = ind_DataFrame[['LT#', 'Project', 'Board', 'Tape', 'Twelve', 
                                     'Ten', 'Nine', 'Eight', 'Prepwork','SQF', 
                                     'Rate', 'Total','Notes']]													
      filepath = f'/content/drive/My Drive/Work_Sheet_{week_of}.xlsx'
      #print(filepath)

      # Add sheet to excel for subsequent employees 

      book = load_workbook(filepath)
      writer = pd.ExcelWriter(filepath, engine='openpyxl')
      writer.book = book
      writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
      print(f'Excel sheet named: {name}_Sheet has been added.')
      sheet_names.append(f'{name}_Sheet')
      ind_DataFrame.to_excel(writer, sheet_name=f'{name}_Sheet', index=False)

      writer.save()

  print('\n Employee sheets created for this week are:')

  for i in sheet_names:
      print(f'        {i}')
  print('Completed.')

# Call function to create weekXentries.xlsx
createWorkSheet(employeeData, employeeTable, 20210101)